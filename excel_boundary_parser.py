"""
Excel Boundary & Details Parser
===============================
Extracts boundary directions (East, West, North, South) directly from the Excel
workbook (Applicants Details / Plot Details) and translates them into legal Hindi.
"""

import openpyxl
import re

def translate_boundary_to_hindi(boundary_text, colony_name="एमराल्ड आश्रय"):
    """
    Translates an English/Hindi boundary entry into standard Hindi deed phrasing.
    Examples:
      - '9 MT. Wide Road' -> '9.00 मीटर चौड़ी सड़क'
      - 'Plot No. L1' -> 'भूखण्ड क्रमांक L1, एमराल्ड आश्रय'
      - 'Plot No. E8' -> 'भूखण्ड क्रमांक ई-8, एमराल्ड आश्रय'
      - 'Plot No. 85 And 86' -> 'भूखण्ड क्रमांक 85 एवं 86, एमराल्ड आश्रय'
      - 'Plot No. 11' -> 'भूखण्ड क्रमांक 11, एमराल्ड गेटवे'
    """
    if not boundary_text:
        return ""
        
    b = str(boundary_text).strip()
    
    # Check for Road patterns (e.g. 9 MT. Wide Road, 9.00 Meter Road, 12M Road)
    road_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:MT\.?|METER|MTR|M)?\s*(?:WIDE)?\s*ROAD', b, re.I)
    if road_match:
        width_val = float(road_match.group(1))
        return f"{width_val:.2f} मीटर चौड़ी सड़क"
        
    # Check for multiple plots like 'Plot No. 85 And 86' or 'Plot No. 85, 86'
    multi_plot_match = re.search(r'PLOT\s*(?:NO\.?)?\s*[-:]?\s*([A-Z0-9\-_/]+)\s*(?:AND|&|\,)\s*([A-Z0-9\-_/]+)', b, re.I)
    if multi_plot_match:
        p1 = multi_plot_match.group(1).strip()
        p2 = multi_plot_match.group(2).strip()
        p1_h = re.sub(r'\b(?:EWS|E)[-_]?(\d+)\b', r'ई-\1', p1, flags=re.I)
        p2_h = re.sub(r'\b(?:EWS|E)[-_]?(\d+)\b', r'ई-\1', p2, flags=re.I)
        return f"भूखण्ड क्रमांक {p1_h} एवं {p2_h}, {colony_name}"

    # Check for single plot number patterns (e.g. Plot No. L1, Plot No. E8, Plot No. 86, Plot No. 11, Plot No. 40, Plot No. E9)
    plot_match = re.search(r'PLOT\s*(?:NO\.?)?\s*[-:]?\s*([A-Z0-9\-_/]+)', b, re.I)
    if plot_match:
        p_num = plot_match.group(1).strip()
        p_num_hindi = re.sub(r'\b(?:EWS|E)[-_]?(\d+)\b', r'ई-\1', p_num, flags=re.I)
        p_num_hindi = re.sub(r'\bL[-_]?(\d+)\b', r'L\1', p_num_hindi, flags=re.I)
        return f"भूखण्ड क्रमांक {p_num_hindi}, {colony_name}"
        
    if re.search(r'GARDEN|PARK', b, re.I):
        return f"गार्डन, {colony_name}"
        
    if re.search(r'OPEN\s*LAND', b, re.I):
        return "खुली भूमि"
        
    return b

def parse_boundaries_from_excel(excel_path):
    """
    Extracts East, West, North, South boundaries directly from the Excel sheet.
    """
    if not excel_path or not openpyxl:
        return {'east': '', 'west': '', 'north': '', 'south': ''}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    colony_name = "एमराल्ड गेटवे"
    app_sheet = None
    for s in wb.sheetnames:
        if 'APPLICANT' in s.upper() or 'PLOT' in s.upper():
            app_sheet = wb[s]
            break
    if not app_sheet:
        app_sheet = wb.active

    for r in range(1, 5):
        val = str(app_sheet.cell(row=r, column=1).value or '')
        if 'AASHRAY' in val.upper():
            colony_name = "एमराल्ड आश्रय"
            break
        elif 'GATEWAY' in val.upper():
            colony_name = "एमराल्ड गेटवे"
            break

    east_raw = ""
    west_raw = ""
    north_raw = ""
    south_raw = ""

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or '').strip()
                val_upper = val.upper()
                if val_upper == 'EAST' and not east_raw:
                    east_raw = str(ws.cell(row=r, column=c+1).value or '').strip()
                elif val_upper == 'WEST' and not west_raw:
                    west_raw = str(ws.cell(row=r, column=c+1).value or '').strip()
                elif val_upper == 'NORTH' and not north_raw:
                    north_raw = str(ws.cell(row=r, column=c+1).value or '').strip()
                elif val_upper == 'SOUTH' and not south_raw:
                    south_raw = str(ws.cell(row=r, column=c+1).value or '').strip()

    return {
        'east': translate_boundary_to_hindi(east_raw, colony_name),
        'west': translate_boundary_to_hindi(west_raw, colony_name),
        'north': translate_boundary_to_hindi(north_raw, colony_name),
        'south': translate_boundary_to_hindi(south_raw, colony_name)
    }
